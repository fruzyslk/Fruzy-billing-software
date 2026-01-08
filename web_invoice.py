# web_invoice.py - Fruzy Web Invoice Generator (Render-compatible)
import os
from flask import Flask, render_template_string, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import re
import json
from datetime import datetime

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# ==== EMBEDDED FRUZY DATA ====
# You can later load this from a JSON if needed
VEGETABLES = [
    {"id": 1, "urdu": "ٹماٹر", "english": "Tomato"},
    {"id": 2, "urdu": "سبز مرچ", "english": "Green Chili"},
    {"id": 3, "urdu": "لہسن", "english": "Garlic"},
    {"id": 4, "urdu": "پیاز", "english": "Onion"},
    {"id": 5, "urdu": "آلو پرانا", "english": "Old Potato"},
]

# In-memory rate list (resets on restart — acceptable for this use case)
rate_list = {}

# ==== HELPER FUNCTIONS ====
def parse_display_item(display_str):
    """Extract English, Urdu, and Size from 'اردو (English) (Large)'"""
    size = 'Normal'
    raw_item = str(display_str).strip()
    m_size = re.search(r"\((Small|Normal|Large)\)\s*$", raw_item, flags=re.IGNORECASE)
    if m_size:
        size = m_size.group(1).capitalize()
        raw_item = re.sub(r"\s*\(%s\)\s*$" % re.escape(m_size.group(1)), '', raw_item, flags=re.IGNORECASE).strip()
    english_name = raw_item
    urdu_name = ''
    m2 = re.search(r"\(([^)]+)\)\s*$", raw_item)
    if m2:
        english_name = m2.group(1).strip()
        urdu_name = raw_item[:m2.start()].strip()
    else:
        english_name = raw_item.strip()
    return english_name, urdu_name, size

def find_urdu_for_english(english_name):
    for v in VEGETABLES:
        if v['english'].strip().lower() == english_name.lower():
            return v['urdu']
    return ''

def match_rate_for_item(item_name):
    english_name, _, _ = parse_display_item(item_name)
    for rate_item, rate in rate_list.items():
        if rate_item.strip().lower() == english_name.lower():
            return rate
    english_name_lower = english_name.lower().strip()
    for rate_item, rate in rate_list.items():
        rate_item_lower = rate_item.lower().strip()
        if english_name_lower in rate_item_lower or rate_item_lower in english_name_lower:
            return rate
    return None

def parse_excel_rate_list(file_content):
    """Parse uploaded Excel and return rate dict"""
    from openpyxl import load_workbook
    from io import BytesIO
    wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    rates = {}
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] and 'Item' in str(row[0]):
                header_found = True
                continue
        if row[0] and row[1]:
            try:
                item_name = str(row[0]).strip()
                rate = float(row[1])
                rates[item_name] = rate
            except (ValueError, TypeError):
                continue
    return rates

# ==== HTML TEMPLATE (Embedded for simplicity) ====
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Fruzy Web Invoice</title>
    <link href="https://fonts.googleapis.com/css2?family=Jameel+Noori+Nastaleeq&display=swap" rel="stylesheet">
    <style>
        body {
            font-family: Arial, sans-serif;
            padding: 20px;
            background-color: #ecf9f2;
        }
        .container { max-width: 1200px; margin: 0 auto; }
        h1 { color: #27ae60; text-align: center; }
        .form-group { margin-bottom: 15px; }
        label { display: inline-block; width: 120px; font-weight: bold; }
        input, select, button { padding: 6px 10px; margin: 2px; }
        .urdu { font-family: 'Jameel Noori Nastaleeq', serif; font-size: 20px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ccc; padding: 8px; text-align: left; }
        th { background-color: #2ecc71; color: white; }
        .total { font-weight: bold; font-size: 18px; margin-top: 15px; }
        .btn { background-color: #2ecc71; color: white; border: none; cursor: pointer; }
        .btn:hover { opacity: 0.9; }
        .rate-status {
            padding: 8px;
            margin: 10px 0;
            border-radius: 4px;
            font-weight: bold;
        }
        .rate-status.success { background-color: #d5f5e3; color: #27ae60; }
        .rate-status.error { background-color: #fadbd8; color: #e74c3c; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🧾 Fruzy Web Invoice Generator</h1>

        <!-- Rate List Upload -->
        <div class="form-group">
            <label>Upload Rate List:</label>
            <form id="rateForm" enctype="multipart/form-data" style="display:inline;">
                <input type="file" name="rate_file" accept=".xlsx,.xls" required>
                <button type="submit" class="btn">Upload</button>
            </form>
            <div id="rateStatus" class="rate-status {{ 'success' if rate_status else 'error' }}">
                {% if rate_status %}✓ {{ rate_status }}{% else %}No rate list loaded{% endif %}
            </div>
        </div>

        <!-- Customer Info -->
        <div class="form-group">
            <label>Customer Name:</label>
            <input type="text" id="customerName" required>
        </div>
        <div class="form-group">
            <label>Phone:</label>
            <input type="text" id="customerPhone">
        </div>

        <!-- Add Item Form -->
        <div class="form-group">
            <label>Item:</label>
            <select id="itemSelect">
                {% for veg in vegetables %}
                    <option value="{{ veg.urdu }} ({{ veg.english }})">{{ veg.urdu }} ({{ veg.english }})</option>
                {% endfor %}
            </select>
            <label>Size:</label>
            <select id="sizeSelect">
                <option>Small</option>
                <option selected>Normal</option>
                <option>Large</option>
            </select>
        </div>
        <div class="form-group">
            <label>Qty:</label>
            <input type="number" id="qtyInput" value="1" step="0.1" min="0.1">
            <label>Unit:</label>
            <select id="unitSelect">
                <option>kg</option>
                <option>piece</option>
                <option>dozen</option>
                <option>bundle</option>
            </select>
            <label>Rate (PKR):</label>
            <input type="number" id="rateInput" value="0" step="0.01" min="0">
            <button class="btn" onclick="addItem()">Add Item</button>
        </div>

        <!-- Items Table -->
        <table id="itemsTable">
            <thead>
                <tr>
                    <th>Item</th>
                    <th>Quantity</th>
                    <th>Rate (PKR)</th>
                    <th>Total (PKR)</th>
                    <th>Action</th>
                </tr>
            </thead>
            <tbody id="itemsBody">
                <!-- Items added via JS -->
            </tbody>
        </table>

        <div class="total">Total: PKR <span id="totalAmount">0.00</span></div>

        <button class="btn" onclick="generateInvoice()" style="margin-top: 20px;">📄 Generate & Download Invoice</button>
    </div>

    <script>
        let items = [];
        const vegetables = {{ vegetables | tojson }};
        const rateList = {{ rate_list | tojson }};

        document.getElementById('itemSelect').addEventListener('change', function() {
            const item = this.value;
            const rate = rateList[item] || rateList[item.split('(')[1].split(')')[0].trim()] || 0;
            document.getElementById('rateInput').value = rate.toFixed(2);
        });

        function addItem() {
            const item = document.getElementById('itemSelect').value;
            const size = document.getElementById('sizeSelect').value;
            const qty = parseFloat(document.getElementById('qtyInput').value) || 0;
            const unit = document.getElementById('unitSelect').value;
            const rate = parseFloat(document.getElementById('rateInput').value) || 0;
            if (qty <= 0 || rate < 0) {
                alert('Invalid quantity or rate');
                return;
            }
            const display = item + ' (' + size + ')';
            const total = qty * rate;
            items.push({display, qty, unit, rate, total});
            renderItems();
            // Reset
            document.getElementById('qtyInput').value = '1';
            document.getElementById('rateInput').value = '0';
        }

        function deleteItem(index) {
            items.splice(index, 1);
            renderItems();
        }

        function renderItems() {
            const body = document.getElementById('itemsBody');
            body.innerHTML = '';
            let total = 0;
            items.forEach((item, idx) => {
                const row = document.createElement('tr');
                row.innerHTML = `
                    <td class="urdu">${item.display}</td>
                    <td>${item.qty} ${item.unit}</td>
                    <td>${item.rate.toFixed(2)}</td>
                    <td>${item.total.toFixed(2)}</td>
                    <td><button onclick="deleteItem(${idx})">🗑️</button></td>
                `;
                body.appendChild(row);
                total += item.total;
            });
            document.getElementById('totalAmount').textContent = total.toFixed(2);
        }

        function generateInvoice() {
            if (items.length === 0) {
                alert('Please add at least one item');
                return;
            }
            if (!document.getElementById('customerName').value.trim()) {
                alert('Please enter customer name');
                return;
            }
            // Submit to backend
            fetch('/generate', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    customer_name: document.getElementById('customerName').value,
                    customer_phone: document.getElementById('customerPhone').value,
                    items: items
                })
            })
            .then(response => response.blob())
            .then(blob => {
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = 'Fruzy_Invoice.xlsx';
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                a.remove();
            })
            .catch(err => alert('Error generating invoice: ' + err));
        }

        // Handle rate list upload
        document.getElementById('rateForm').addEventListener('submit', function(e) {
            e.preventDefault();
            const formData = new FormData(this);
            fetch('/upload_rate', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    location.reload(); // Refresh to show updated rate status
                } else {
                    alert('Upload failed: ' + data.error);
                }
            });
        });
    </script>
</body>
</html>
'''

# ==== ROUTES ====
@app.route('/')
def index():
    global rate_list
    rate_status = f"Rate list loaded: {len(rate_list)} items" if rate_list else None
    return render_template_string(
        HTML_TEMPLATE,
        vegetables=VEGETABLES,
        rate_list=rate_list,
        rate_status=rate_status
    )

@app.route('/upload_rate', methods=['POST'])
def upload_rate():
    global rate_list
    try:
        if 'rate_file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        file = request.files['rate_file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            return jsonify({'success': False, 'error': 'Invalid file type'}), 400
        file_content = file.read()
        rate_list = parse_excel_rate_list(file_content)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/generate', methods=['POST'])
def generate_invoice():
    try:
        data = request.get_json()
        customer_name = data.get('customer_name', '').strip()
        customer_phone = data.get('customer_phone', '').strip()
        items = data.get('items', [])
        if not customer_name:
            return 'Customer name required', 400
        if not items:
            return 'No items provided', 400

        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "FRUZY"
        ws['A1'].font = Font(name='Calibri', size=72, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f"{col}1"].fill = header_fill
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        invoice_num = 9999  # Placeholder — no counter in web version
        ws['B2'] = f"Invoice #{invoice_num}"
        ws['B2'].font = Font(name='Calibri', size=11)
        ws['E2'] = datetime.now().strftime("%d-%b-%Y %I:%M %p")
        ws['E2'].font = Font(name='Calibri', size=11)
        ws['E2'].alignment = Alignment(horizontal='right')

        ws['B4'] = "Customer Name:"
        ws['B4'].font = Font(name='Calibri', size=11)
        ws['C4'] = customer_name
        ws['C4'].font = Font(name='Calibri', size=11, bold=True)
        ws['E4'] = customer_phone
        ws['E4'].font = Font(name='Calibri', size=11, bold=True)

        # Headers
        header_row = 6
        headers = ['No.', 'Item Name', 'Quantity', 'Rate (PKR)', 'Total (PKR)']
        header_fill2 = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill2
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Items
        row = header_row + 1
        total = 0.0
        for idx, item in enumerate(items, 1):
            raw_item = item['display']
            english_name, urdu_name, size = parse_display_item(raw_item)
            if not urdu_name:
                urdu_name = find_urdu_for_english(english_name)
            urdu_size_map = {'Small': 'چھوٹا سائز', 'Normal': 'درمیانہ سائز', 'Large': 'بڑا سائز'}
            english_size_map = {'Small': 'small size', 'Large': 'big size'}
            english_plain = english_name
            if size.lower() != 'normal':
                english_plain = f"{english_name} {english_size_map.get(size, size)}"
            urdu_cell = urdu_name
            if urdu_name and size.lower() != 'normal':
                urdu_cell = f"{urdu_name} {urdu_size_map.get(size, '')}"
            if urdu_name:
                item_cell_value = f"{urdu_cell} ({english_plain})"
            else:
                item_cell_value = english_plain

            ws.cell(row=row, column=1, value=idx).border = border
            cell = ws.cell(row=row, column=2, value=item_cell_value)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = border
            ws.cell(row=row, column=3, value=f"{item['qty']} {item['unit']}").border = border
            rate_cell = ws.cell(row=row, column=4, value=item['rate'])
            rate_cell.number_format = '#,##0.00'
            rate_cell.border = border
            tot_cell = ws.cell(row=row, column=5, value=item['total'])
            tot_cell.number_format = '#,##0.00'
            tot_cell.border = border
            total += item['total']
            row += 1

        # Total
        ws.cell(row=row + 1, column=1, value="Total Amount:").font = Font(name='Calibri', size=11, bold=True)
        total_cell = ws.cell(row=row + 1, column=5, value=total)
        total_cell.font = Font(name='Calibri', size=11, bold=True)
        total_cell.number_format = '#,##0.00'

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Fruzy_Invoice.xlsx'
        )
    except Exception as e:
        return f'Error: {str(e)}', 500

# ==== RUN ====
if __name__ == '__main__':
    app.run(debug=True)
